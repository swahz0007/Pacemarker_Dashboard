"""
Excel 文件处理器模块
支持 .xls 和 .xlsx 格式

智能 Sheet 选择策略:
  1. 从文件名提取患者姓名
  2. 扫描所有有数据的 sheet，找包含该姓名的 sheet
  3. 如果找到匹配的 sheet → 使用它
  4. 如果没找到 → 回退到最后一个有数据的 sheet
"""

import logging
import openpyxl
from core.utils import extract_name_from_filename_legacy

logger = logging.getLogger(__name__)
try:
    import xlrd
except ImportError:  # pragma: no cover - depends on runtime environment
    xlrd = None


def _extract_name_from_path(filepath):
    """从文件路径提取患者姓名（轻量版，避免循环导入）"""
    return extract_name_from_filename_legacy(filepath)


class XlsHandler:
    """处理旧版 .xls 格式文件"""

    def __init__(self, filepath):
        if xlrd is None:
            raise RuntimeError("xlrd is not installed; .xls files cannot be processed in this environment")
        self.book = xlrd.open_workbook(filepath, formatting_info=True)
        patient_name = _extract_name_from_path(filepath)
        self.sheet = self._select_best_sheet(patient_name)
        self.nrows = self.sheet.nrows
        self.ncols = self.sheet.ncols

    def _select_best_sheet(self, patient_name):
        """智能选择 Sheet：优先匹配患者姓名，否则选最后一个有数据的"""
        best_sheet = self.book.sheet_by_index(0)
        candidates = []

        for i in range(self.book.nsheets):
            s = self.book.sheet_by_index(i)
            if s.nrows <= 2:
                continue  # 空 sheet
            candidates.append(s)

            # 在前 5 行中查找患者姓名
            if patient_name and len(patient_name) >= 2:
                for r in range(min(5, s.nrows)):
                    for c in range(min(10, s.ncols)):
                        v = str(s.cell_value(r, c)).strip()
                        if v and patient_name in v:
                            return s  # 找到匹配的 sheet

        # 没有匹配 → 选最后一个有数据的
        return candidates[-1] if candidates else best_sheet

    def get_cell_value(self, r, c):
        return self.sheet.cell_value(r, c) if r < self.nrows and c < self.ncols else ""

    def is_blue_cell(self, r, c):
        if r >= self.nrows or c >= self.ncols:
            return False
        try:
            xf_idx = self.sheet.cell_xf_index(r, c)
            return self.book.xf_list[xf_idx].background.pattern_colour_index == 31
        except IndexError:
            return False
        except Exception as e:
            logger.warning(f"检查蓝色单元格失败 ({r},{c}): {e}")
            return False

    def close(self):
        """关闭工作簿，释放底层资源"""
        if not self.book:
            return
        try:
            self.book.release_resources()
        except AttributeError:
            logger.debug("xlrd book does not support release_resources")


class XlsxHandler:
    """处理新版 .xlsx 格式文件"""

    def __init__(self, filepath):
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        patient_name = _extract_name_from_path(filepath)
        self.sheet = self._select_best_sheet(patient_name)
        self.nrows = self.sheet.max_row
        self.ncols = self.sheet.max_column

    def _select_best_sheet(self, patient_name):
        """智能选择 Sheet：优先匹配患者姓名"""
        candidates = []
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            if not ws.max_row or ws.max_row <= 2:
                continue
            candidates.append(ws)

            # 在前 5 行查找患者姓名
            if patient_name and len(patient_name) >= 2:
                for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row),
                                         max_col=10, values_only=True):
                    for v in row:
                        if v and patient_name in str(v).strip():
                            return ws

        # 回退：最后一个有数据的 sheet
        return candidates[-1] if candidates else self.wb.active

    def get_cell_value(self, r, c):
        try:
            return self.sheet.cell(row=r + 1, column=c + 1).value
        except IndexError:
            return ""
        except Exception as e:
            logger.warning(f"获取单元格值失败 ({r},{c}): {e}")
            return ""

    def is_blue_cell(self, r, c):
        try:
            cell = self.sheet.cell(row=r + 1, column=c + 1)
            return cell.fill.fgColor.type == "theme" and cell.fill.fgColor.theme == 4
        except AttributeError:
            return False
        except IndexError:
            return False
        except Exception as e:
            logger.warning(f"检查蓝色单元格失败 ({r},{c}): {e}")
            return False

    def close(self):
        """关闭 Workbook 释放资源"""
        if self.wb:
            self.wb.close()


def get_handler(filepath):
    """根据文件扩展名返回对应的处理器"""
    if filepath.lower().endswith(".xls"):
        return XlsHandler(filepath)
    return XlsxHandler(filepath)


def is_xls_supported():
    """当前运行环境是否支持 .xls 读取"""
    return xlrd is not None
